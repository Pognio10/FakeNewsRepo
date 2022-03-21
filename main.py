#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from who import WHO
from when import WHEN
from where import WHERE 
from what import WHAT
from why import WHY
from searchgoogle import searchG
from count_upper import COUNT_UPPER
from ScrapSel import ImageAnalize
from checkurl import validate_url
import pandas as pd
import openpyxl
import os
from pathlib import Path
import shutil
from langdetect import detect
from deep_translator import GoogleTranslator



df_test = pd.DataFrame([], columns=["Url","Score_Who","Score_When","Score_What","Score_Where","Score_Why","Score_Count_upper","Score_SearchText","Score_Images","Score_Tot","Orig_Label"])
df_test.head()



# book = openpyxl.load_workbook('link_site.xlsx')
# sheet = book.active

# # for i in range(1,(len(sheet['A']))):
# for i in range(22,23):
# 	a3 = sheet.cell(row=i, column=1)
# 	label=sheet.cell(row=i, column=2)
# 	url=a3.value
# 	labels=label.value

url = input("Please enter the url to analyze: ")
url = validate_url(url)

score_tot=0


text, auth, scoreWho, title, image = WHO(url)
print("\n", auth)
print("The score of WHO is " +str(scoreWho), "\n")
score_tot=score_tot+scoreWho


lang = detect(text)
if lang == 'en':
	text = text
else:
	to_translate = text
	translated = GoogleTranslator(source='auto', target='en').translate(to_translate)
	text = translated


days_diff, scoreWhen=WHEN(text)
print("The score of WHEN is " + str(scoreWhen), "\n")
score_tot=score_tot+scoreWhen

scoreWhat=WHAT()
print("The score of WHAT is " + str(scoreWhat), "\n")
score_tot=score_tot+scoreWhat


distance, scoreWhere=WHERE(text)
print("The distance between your location and the place where the events took place is: ", distance, "KM", "\n")
print("The score of WHERE is " + str(scoreWhere), "\n")
score_tot=score_tot+scoreWhere
# scoreWhere = 1.0
# score_tot = score_tot + scoreWhere

scoreWhy=WHY()
print("The score of WHY is " + str(scoreWhy), "\n")
score_tot=score_tot+scoreWhy

print(title)

rep, scoreCount=COUNT_UPPER(title)
print("The score of HOW1 is " + str(scoreCount), "\n")
score_tot=score_tot+scoreCount

search, scoreSearchText=searchG(title)
print("The score of HOW2 is " + str(scoreSearchText), "\n")
score_tot=score_tot+scoreSearchText

scoreImage = ImageAnalize()
print("The score of HOW3 is " + str(scoreImage), "\n")
score_tot=score_tot+scoreImage



print("The news has a reliability value of " + str(score_tot) + "/8.7")



# arrayArt=[url,scoreWho,scoreWhen,scoreWhat,scoreWhere,scoreWhy,scoreCount,scoreSearchText,scoreImage,score_tot,labels]
# df_testS = pd.DataFrame([arrayArt], columns=["Url","Score_Who","Score_When","Score_What","Score_Where","Score_Why","Score_Count_upper","Score_SearchText","Score_Images","Score_Tot","Orig_Label"])
# df_test=df_test.append(df_testS)
# print(df_test)


	

if score_tot > 5.4:
	print("The News is probably a Real News")
else:
	print("The news is probably a Fake news")


# if not os.path.isdir("TestFile/"):
#     os.mkdir("TestFile/")
# else:
#     dir_path=Path("TestFile/")
#     shutil.rmtree(dir_path)
#     os.mkdir("TestFile/")

# df_test.to_excel("TestFile/testFile.xlsx",index=True)  

